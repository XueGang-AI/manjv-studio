#!/usr/bin/env python3
"""解析 XLSX 文件 → 输出结构化 JSON"""
import json, os
from openpyxl import load_workbook

FILES = [
    ("【先看这个】分镜画面提示词.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/分镜提示词整理/【先看这个】分镜画面提示词.xlsx",
     "分镜提示词整理"),
    ("即梦100多组神级指令合集.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/分镜提示词整理/即梦100多组（700+个）神级指令合集.xlsx",
     "分镜提示词整理"),
    ("AI视频脚本分镜模板_共300条.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/分镜提示词整理/AI视频脚本分镜模板_共300条.xlsx",
     "分镜提示词整理"),
    ("15种ai漫剧题材的人物提示词案例.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/小说漫剧提示词/15种ai漫剧题材的人物提示词案例.xlsx",
     "小说漫剧提示词"),
    ("AI生成人物角色提示词通用公式表格模板.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/小说漫剧提示词/AI生成人物角色提示词通用公式表格模板.xlsx",
     "小说漫剧提示词"),
    ("即梦4.5 AI 漫剧运镜提示词50条.xlsx",
     "/Users/xuegang/Desktop/My Project/AI漫剧制作/分镜提示词＋AI指令整理/小说漫剧提示词/即梦 4.5 AI 漫剧运镜提示词 50 条.xlsx",
     "小说漫剧提示词"),
]

OUTPUT = "/Users/xuegang/Desktop/My Project/manjv-studio/scripts/output/xlsx_parsed.json"
results = []

for name, path, source_dir in FILES:
    entry = {"file": name, "source_dir": source_dir}
    if not os.path.exists(path):
        entry["status"] = "NOT_FOUND"
        entry["path"] = path
        results.append(entry)
        print(f"❌ {name}: NOT FOUND at {path}")
        continue
    
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_vals = [str(c) if c is not None else "" for c in row]
                if i == 0:
                    headers = row_vals
                else:
                    if any(v.strip() for v in row_vals):
                        rows.append(dict(zip(headers, row_vals)))
            sheets_data[sheet_name] = {
                "headers": headers,
                "row_count": len(rows),
                "rows": rows
            }
        
        total_rows = sum(s["row_count"] for s in sheets_data.values())
        entry["status"] = "OK"
        entry["sheets"] = list(sheets_data.keys())
        entry["total_rows"] = total_rows
        entry["data"] = sheets_data
        results.append(entry)
        print(f"✅ {name}: {len(sheets_data)} sheets, {total_rows} rows")
        
    except Exception as e:
        entry["status"] = "ERROR"
        entry["error"] = str(e)
        results.append(entry)
        print(f"❌ {name}: ERROR - {e}")

with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print(f"\n📄 Output: {OUTPUT}")
